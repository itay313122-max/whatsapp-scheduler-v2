"""
build_citizen_impact.py — builds citizen_impact_full.xlsx with 4 RTL sheets.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1A3A5C"
MED_BLUE    = "2563EB"
LIGHT_BLUE  = "DBEAFE"
GOLD        = "F59E0B"
GREEN       = "10B981"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F1F5F9"
DARK_TEXT   = "1E293B"

# ── Shared style helpers ──────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=DARK_TEXT, size=10, name="Arial") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)

def _border() -> Border:
    thin = Side(style="thin", color="B0BEC5")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(horizontal="right", vertical="center", wrap=True) -> Alignment:
    return Alignment(
        horizontal=horizontal, vertical=vertical,
        wrap_text=wrap, readingOrder=2  # RTL
    )

def _hdr_style(cell, text, bg=DARK_BLUE, fg=WHITE, size=11, bold=True):
    cell.value = text
    cell.font = _font(bold=bold, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.alignment = _align("center")
    cell.border = _border()

def _col_hdr(cell, text, bg=MED_BLUE, fg=WHITE):
    cell.value = text
    cell.font = _font(bold=True, color=fg, size=10)
    cell.fill = _fill(bg)
    cell.alignment = _align("center")
    cell.border = _border()

def _data_cell(cell, value, row_idx: int, bold=False, number_format=None, align="right"):
    cell.value = value
    bg = LIGHT_BLUE if row_idx % 2 == 0 else WHITE
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, size=10)
    cell.alignment = _align(align)
    cell.border = _border()
    if number_format:
        cell.number_format = number_format

def _set_rtl(ws):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = True

def _set_col_width(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

def _set_row_height(ws, row_heights: dict):
    for row_num, height in row_heights.items():
        ws.row_dimensions[row_num].height = height

def _merge_title(ws, cell_range, text, bg=DARK_BLUE, fg=WHITE, size=14):
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = text
    top_left.font = _font(bold=True, color=fg, size=size)
    top_left.fill = _fill(bg)
    top_left.alignment = _align("center")
    top_left.border = _border()

def _section_title(ws, row, col_start, col_end, text, bg=GOLD):
    col_s = get_column_letter(col_start)
    col_e = get_column_letter(col_end)
    ws.merge_cells(f"{col_s}{row}:{col_e}{row}")
    cell = ws[f"{col_s}{row}"]
    cell.value = text
    cell.font = _font(bold=True, color=WHITE, size=11)
    cell.fill = _fill(bg)
    cell.alignment = _align("center")
    cell.border = _border()
    ws.row_dimensions[row].height = 22

# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 1 — סיכום כללי
# ══════════════════════════════════════════════════════════════════════════════
def build_sheet1(wb: Workbook):
    ws = wb.create_sheet("📋 סיכום כללי")
    _set_rtl(ws)

    _set_col_width(ws, {
        "A": 30, "B": 45, "C": 45, "D": 20
    })

    # Main title
    ws.merge_cells("A1:D1")
    _hdr_style(ws["A1"], "Citizen Impact — סיטיזן אימפקט | סיכום כללי",
               bg=DARK_BLUE, fg=WHITE, size=16)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    _hdr_style(ws["A2"], "פלטפורמת AI לחוסן פיננסי | FinTech | תל אביב, 2021",
               bg=MED_BLUE, fg=WHITE, size=12)
    ws.row_dimensions[2].height = 24

    # ── Section: פרטי חברה ──
    _section_title(ws, 4, 1, 4, "🏢  פרטי החברה")

    headers_info = [
        ("שדה", "ערך / פרטים"),
    ]
    _col_hdr(ws["A5"], "שדה")
    _col_hdr(ws["B5"], "ערך / פרטים")
    ws.merge_cells("B5:D5")
    ws.row_dimensions[5].height = 20

    company_data = [
        ("שם החברה",           "Citizen Impact / סיטיזן אימפקט"),
        ("תחום",               "FinTech | AI פיננסי | חוסן פיננסי"),
        ("שנת הקמה",           "2021"),
        ("מיקום",              "תל אביב, ישראל"),
        ("אתר B2B",            "www.citizens-ai.com"),
        ("פלטפורמת לקוחות",    "landing.is.citizen-ai.org"),
        ("אימייל",             "info@citizens-ai.com"),
        ("טלפון",              "074-7281259"),
    ]
    for i, (field, value) in enumerate(company_data):
        row = 6 + i
        _data_cell(ws[f"A{row}"], field, i, bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        _data_cell(ws[f"B{row}"], value, i)
        ws.row_dimensions[row].height = 18

    # ── Section: חזון ומשימה ──
    vision_row = 6 + len(company_data) + 1
    _section_title(ws, vision_row, 1, 4, "🎯  חזון ומשימה")

    vision_data = [
        ("חזון",     "להפוך חוסן פיננסי לנגיש לכל אזרח — בכל שלב בחיים"),
        ("משימה",    "מערכת AI שמזהה, מנבאת ומנחה משפחות לשיקום פיננסי אמיתי"),
        ("ייחודיות", "שילוב פסיכולוגיה פיננסית, AI ונתוני בנקים בזמן אמת"),
        ("קהל יעד",  "אזרחים פרטיים, מוסדות פיננסיים, עיריות ורשויות ממשל"),
    ]
    for i, (field, value) in enumerate(vision_data):
        row = vision_row + 1 + i
        _data_cell(ws[f"A{row}"], field, i, bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        _data_cell(ws[f"B{row}"], value, i)
        ws.row_dimensions[row].height = 22

    # ── Section: KPIs מרכזיים (ערכי בסיס לנוסחאות) ──
    kpi_base_row = vision_row + 1 + len(vision_data) + 1
    _section_title(ws, kpi_base_row, 1, 4, "📊  KPIs מרכזיים")

    ws.row_dimensions[kpi_base_row].height = 22
    kpi_header_row = kpi_base_row + 1
    _col_hdr(ws[f"A{kpi_header_row}"], "מדד")
    _col_hdr(ws[f"B{kpi_header_row}"], "ערך בסיס (%)")
    _col_hdr(ws[f"C{kpi_header_row}"], "אחוז (נוסחה)")
    _col_hdr(ws[f"D{kpi_header_row}"], "סטטוס")
    ws.row_dimensions[kpi_header_row].height = 20

    # Base values for formulas (stored as numbers, percentages computed via formula)
    kpi_data = [
        ("גידול בהכנסות — משתמשים שהשיגו",   80,  "מוכח"),
        ("גידול בהכנסות — שיעור שיפור",        20,  "מוכח"),
        ("מגירעון ליתרת זכות",                  75,  "מוכח"),
        ("כשירות אשראי",                        90,  "מוכח"),
    ]
    kpi_start = kpi_header_row + 1
    for i, (label, base_val, status) in enumerate(kpi_data):
        row = kpi_start + i
        b_col = f"B{row}"
        c_col = f"C{row}"
        _data_cell(ws[f"A{row}"], label, i, bold=True)
        _data_cell(ws[b_col], base_val, i)
        ws[b_col].number_format = "0"
        # Formula: percentage representation
        ws[c_col].value = f"={b_col}/100"
        ws[c_col].number_format = FORMAT_PERCENTAGE
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws[c_col].fill = _fill(bg)
        ws[c_col].font = _font(bold=True, color=GREEN, size=10)
        ws[c_col].alignment = _align("center")
        ws[c_col].border = _border()
        _data_cell(ws[f"D{row}"], status, i, bold=False)
        ws.row_dimensions[row].height = 18

    # Summary row with formula
    summary_row = kpi_start + len(kpi_data)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    cell = ws[f"A{summary_row}"]
    cell.value = "ממוצע KPIs (נוסחה)"
    cell.font = _font(bold=True, color=WHITE, size=10)
    cell.fill = _fill(DARK_BLUE)
    cell.alignment = _align("center")
    cell.border = _border()

    avg_cell = ws[f"C{summary_row}"]
    avg_cell.value = f"=AVERAGE(B{kpi_start}:B{kpi_start + len(kpi_data) - 1})/100"
    avg_cell.number_format = FORMAT_PERCENTAGE
    avg_cell.font = _font(bold=True, color=WHITE, size=10)
    avg_cell.fill = _fill(DARK_BLUE)
    avg_cell.alignment = _align("center")
    avg_cell.border = _border()

    ws[f"D{summary_row}"].value = ""
    ws[f"D{summary_row}"].fill = _fill(DARK_BLUE)
    ws[f"D{summary_row}"].border = _border()
    ws.row_dimensions[summary_row].height = 22

    # Freeze panes
    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 2 — מוצר וטכנולוגיה
# ══════════════════════════════════════════════════════════════════════════════
def build_sheet2(wb: Workbook):
    ws = wb.create_sheet("🛠️ מוצר וטכנולוגיה")
    _set_rtl(ws)

    _set_col_width(ws, {
        "A": 8, "B": 35, "C": 55, "D": 25
    })

    # Main title
    ws.merge_cells("A1:D1")
    _hdr_style(ws["A1"], "מוצר וטכנולוגיה — 5 שלבי המערכת",
               bg=DARK_BLUE, fg=WHITE, size=16)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    _hdr_style(ws["A2"], "פלטפורמת AI לחוסן פיננסי | זיהוי · חיזוי · ליווי · אוטומציה · ניהול",
               bg=MED_BLUE, fg=WHITE, size=11)
    ws.row_dimensions[2].height = 24

    # ── 5 Steps ──
    _section_title(ws, 4, 1, 4, "🔄  5 שלבי הפלטפורמה")

    _col_hdr(ws["A5"], "שלב")
    _col_hdr(ws["B5"], "שם השלב")
    _col_hdr(ws["C5"], "תיאור")
    _col_hdr(ws["D5"], "טכנולוגיה")
    ws.row_dimensions[5].height = 20

    steps = [
        ("1", "זיהוי מסלול ליווי",
         "מזהה משפחות בסיכון פיננסי בזמן אמת על בסיס נתוני בנק",
         "Real-time AI | Data Analytics"),
        ("2", "חיזוי פוטנציאל צמיחה",
         "מנוע AI מנבא ומייצר תוכנית הבראה פיננסית מותאמת אישית",
         "ML Prediction | NLP"),
        ("3", "תוכנית ליווי פרסונלית",
         "מפת שיקום עם צעדים ברורים, ציר זמן ויעדים מדידים",
         "Personalization Engine"),
        ("4", "אוטומציה של המלצות",
         "התראות AI בזמן אמת, ללא תיווך ידני — ישירות ללקוח",
         "Automation | Push AI Alerts"),
        ("5", "ניהול ויישום",
         "מנטור תוצאות, מעקב KPI, דוחות — עד מימוש מלא",
         "Dashboard | KPI Tracker"),
    ]
    for i, (step, name, desc, tech) in enumerate(steps):
        row = 6 + i
        _data_cell(ws[f"A{row}"], step, i, bold=True, align="center")
        _data_cell(ws[f"B{row}"], name, i, bold=True)
        _data_cell(ws[f"C{row}"], desc, i)
        _data_cell(ws[f"D{row}"], tech, i)
        ws.row_dimensions[row].height = 40

    # ── KPIs ──
    kpi_title_row = 12
    _section_title(ws, kpi_title_row, 1, 4, "📊  KPIs מוכחים — ערכי בסיס")

    kpi_hdr_row = kpi_title_row + 1
    _col_hdr(ws[f"A{kpi_hdr_row}"], "מזהה")
    _col_hdr(ws[f"B{kpi_hdr_row}"], "מדד")
    _col_hdr(ws[f"C{kpi_hdr_row}"], "שיעור (% — בסיס)")
    _col_hdr(ws[f"D{kpi_hdr_row}"], "חישוב אחוז")
    ws.row_dimensions[kpi_hdr_row].height = 20

    kpi_rows = [
        ("K1", "גידול הכנסות — אחוז משתמשים שהשיגו", 80),
        ("K2", "גידול הכנסות — שיעור שיפור ממוצע",   20),
        ("K3", "מגירעון ליתרת זכות",                   75),
        ("K4", "כשירות אשראי",                         90),
    ]
    kpi_data_start = kpi_hdr_row + 1
    for i, (kid, label, pct) in enumerate(kpi_rows):
        row = kpi_data_start + i
        c_col = f"C{row}"
        d_col = f"D{row}"
        _data_cell(ws[f"A{row}"], kid, i, bold=True, align="center")
        _data_cell(ws[f"B{row}"], label, i, bold=True)
        _data_cell(ws[c_col], pct, i)
        ws[c_col].number_format = "0\"%\""
        # Formula for percentage column
        ws[d_col].value = f"={c_col}/100"
        ws[d_col].number_format = FORMAT_PERCENTAGE
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws[d_col].fill = _fill(bg)
        ws[d_col].font = _font(bold=True, color=GREEN)
        ws[d_col].alignment = _align("center")
        ws[d_col].border = _border()
        ws.row_dimensions[row].height = 20

    # Average row
    avg_row = kpi_data_start + len(kpi_rows)
    ws.merge_cells(f"A{avg_row}:B{avg_row}")
    ws[f"A{avg_row}"].value = "ממוצע כלל KPIs (נוסחה)"
    ws[f"A{avg_row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{avg_row}"].fill = _fill(DARK_BLUE)
    ws[f"A{avg_row}"].alignment = _align("center")
    ws[f"A{avg_row}"].border = _border()

    ws[f"C{avg_row}"].value = f"=AVERAGE(C{kpi_data_start}:C{avg_row - 1})"
    ws[f"C{avg_row}"].number_format = "0.0\"%\""
    ws[f"C{avg_row}"].font = _font(bold=True, color=WHITE)
    ws[f"C{avg_row}"].fill = _fill(DARK_BLUE)
    ws[f"C{avg_row}"].alignment = _align("center")
    ws[f"C{avg_row}"].border = _border()

    ws[f"D{avg_row}"].value = f"=C{avg_row}/100"
    ws[f"D{avg_row}"].number_format = FORMAT_PERCENTAGE
    ws[f"D{avg_row}"].font = _font(bold=True, color=WHITE)
    ws[f"D{avg_row}"].fill = _fill(DARK_BLUE)
    ws[f"D{avg_row}"].alignment = _align("center")
    ws[f"D{avg_row}"].border = _border()
    ws.row_dimensions[avg_row].height = 22

    # ── Active pages ──
    pages_title_row = avg_row + 2
    _section_title(ws, pages_title_row, 1, 4, "🌐  דפי לקוחות פעילים")

    pages_hdr_row = pages_title_row + 1
    _col_hdr(ws[f"A{pages_hdr_row}"], "#")
    _col_hdr(ws[f"B{pages_hdr_row}"], "שם הדף")
    _col_hdr(ws[f"C{pages_hdr_row}"], "URL")
    _col_hdr(ws[f"D{pages_hdr_row}"], "סטטוס")
    ws.row_dimensions[pages_hdr_row].height = 20

    pages = [
        ("CustomerHome35",        "client.is.citizen-ai.org/#/?customer=35", "פעיל"),
        ("CustomerHome36",        "client.is.citizen-ai.org/#/?customer=36", "פעיל"),
        ("CustomerHome37",        "client.is.citizen-ai.org/#/?customer=37", "פעיל"),
        ("CustomerHome42",        "client.is.citizen-ai.org/#/?customer=42", "פעיל"),
        ("CustomerHome44",        "landing.is.citizen-ai.org/#/?customer=44", "פעיל"),
        ("CustomerHome45",        "landing.is.citizen-ai.org/#/?customer=45", "פעיל"),
        ("CustomerHome46",        "landing.is.citizen-ai.org/#/?customer=46", "פעיל"),
        ("CustomerHome47",        "landing.is.citizen-ai.org/#/?customer=47", "פעיל"),
        ("CustomerHome48",        "landing.is.citizen-ai.org/#/?customer=48", "פעיל"),
        ("PrivacyPolicy_47",      "landing.is.citizen-ai.org/privacy/47",     "פעיל"),
        ("TermsOfService_47",     "landing.is.citizen-ai.org/terms/47",       "פעיל"),
        ("PrivacyPolicy_36/46/48","landing.is.citizen-ai.org/privacy/...",    "פעיל"),
    ]
    pages_data_start = pages_hdr_row + 1
    for i, (name, url, status) in enumerate(pages):
        row = pages_data_start + i
        _data_cell(ws[f"A{row}"], i + 1, i, align="center")
        _data_cell(ws[f"B{row}"], name, i, bold=True)
        _data_cell(ws[f"C{row}"], url, i)
        _data_cell(ws[f"D{row}"], status, i, bold=True)
        # Color status green
        ws[f"D{row}"].font = _font(bold=True, color=GREEN)
        ws.row_dimensions[row].height = 18

    # Count formula row
    count_row = pages_data_start + len(pages)
    ws.merge_cells(f"A{count_row}:B{count_row}")
    ws[f"A{count_row}"].value = "סה\"כ דפים פעילים (נוסחה)"
    ws[f"A{count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{count_row}"].fill = _fill(DARK_BLUE)
    ws[f"A{count_row}"].alignment = _align("center")
    ws[f"A{count_row}"].border = _border()

    ws[f"C{count_row}"].value = f"=COUNTA(B{pages_data_start}:B{count_row - 1})"
    ws[f"C{count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"C{count_row}"].fill = _fill(DARK_BLUE)
    ws[f"C{count_row}"].alignment = _align("center")
    ws[f"C{count_row}"].border = _border()

    ws[f"D{count_row}"].value = "דפים"
    ws[f"D{count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"D{count_row}"].fill = _fill(DARK_BLUE)
    ws[f"D{count_row}"].alignment = _align("center")
    ws[f"D{count_row}"].border = _border()
    ws.row_dimensions[count_row].height = 22

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 3 — צוות ומשקיעים
# ══════════════════════════════════════════════════════════════════════════════
def build_sheet3(wb: Workbook):
    ws = wb.create_sheet("👥 צוות ומשקיעים")
    _set_rtl(ws)

    _set_col_width(ws, {
        "A": 28, "B": 30, "C": 38, "D": 20
    })

    ws.merge_cells("A1:D1")
    _hdr_style(ws["A1"], "צוות מייסדים ומשקיעים — Citizen Impact",
               bg=DARK_BLUE, fg=WHITE, size=16)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    _hdr_style(ws["A2"], "הנהלה בכירה · משקיעים אסטרטגיים · שותפים",
               bg=MED_BLUE, fg=WHITE, size=11)
    ws.row_dimensions[2].height = 24

    # ── Team ──
    _section_title(ws, 4, 1, 4, "👤  צוות מייסדים והנהלה")

    _col_hdr(ws["A5"], "שם")
    _col_hdr(ws["B5"], "תפקיד")
    _col_hdr(ws["C5"], "התמחות")
    _col_hdr(ws["D5"], "ניסיון")
    ws.row_dimensions[5].height = 20

    team = [
        ('ד"ר דגנית בידרמן', "מייסדת ומנכ\"לית", "פסיכולוגיה פיננסית, AI",     "15+ שנה"),
        ("ינון ברוש",          "מייסד ו-CTO",       "מערכות AI, ארכיטקטורת ענן",  "20 שנה"),
        ("שירה רדובן",         "CFO",                "סיכוני אשראי, גיוס הון",      "ניסיון רב"),
        ("ישראל דנצינגר",      "יו\"ר הדירקטוריון", "היי-טק בינלאומי, ייעוץ",     "בכיר"),
    ]
    team_data_start = 6
    for i, (name, role, spec, exp) in enumerate(team):
        row = team_data_start + i
        _data_cell(ws[f"A{row}"], name, i, bold=True)
        _data_cell(ws[f"B{row}"], role, i)
        _data_cell(ws[f"C{row}"], spec, i)
        _data_cell(ws[f"D{row}"], exp, i)
        ws.row_dimensions[row].height = 28

    # Count formula
    team_count_row = team_data_start + len(team)
    ws.merge_cells(f"A{team_count_row}:C{team_count_row}")
    ws[f"A{team_count_row}"].value = "סה\"כ חברי הנהלה (נוסחה)"
    ws[f"A{team_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{team_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"A{team_count_row}"].alignment = _align("center")
    ws[f"A{team_count_row}"].border = _border()
    ws[f"D{team_count_row}"].value = f"=COUNTA(A{team_data_start}:A{team_count_row - 1})"
    ws[f"D{team_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"D{team_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"D{team_count_row}"].alignment = _align("center")
    ws[f"D{team_count_row}"].border = _border()
    ws.row_dimensions[team_count_row].height = 22

    # ── Investors & Partners ──
    inv_title_row = team_count_row + 2
    _section_title(ws, inv_title_row, 1, 4, "💼  משקיעים ושותפים אסטרטגיים")

    inv_hdr_row = inv_title_row + 1
    _col_hdr(ws[f"A{inv_hdr_row}"], "שם / גוף")
    _col_hdr(ws[f"B{inv_hdr_row}"], "תפקיד / קשר")
    ws.merge_cells(f"B{inv_hdr_row}:D{inv_hdr_row}")
    ws.row_dimensions[inv_hdr_row].height = 20

    investors = [
        ("Mastercard Innovation Lab", "שותף אסטרטגי"),
        ("גליה אלבין",                "משקיעה (לשעבר מנכ\"לית בנק לאומי)"),
        ("חמי פרס",                   "משקיע (לשעבר מנכ\"ל Pitango)"),
        ("קארון בילסקי",              "מייסדת BSP Funds Ltd"),
        ("משרד שבלת",                 "השקעת אימפקט + ייעוץ משפטי"),
        ("רשות החדשנות",              "שותף ממשלתי"),
        ("רשות שוק ההון",             "שותף רגולטורי"),
        ("עוגן",                       "שותף — פרויקט נפגעי מלחמה"),
    ]
    inv_data_start = inv_hdr_row + 1
    for i, (name, role) in enumerate(investors):
        row = inv_data_start + i
        _data_cell(ws[f"A{row}"], name, i, bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        _data_cell(ws[f"B{row}"], role, i)
        ws.row_dimensions[row].height = 22

    # Count formula
    inv_count_row = inv_data_start + len(investors)
    ws.merge_cells(f"A{inv_count_row}:C{inv_count_row}")
    ws[f"A{inv_count_row}"].value = "סה\"כ משקיעים/שותפים (נוסחה)"
    ws[f"A{inv_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{inv_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"A{inv_count_row}"].alignment = _align("center")
    ws[f"A{inv_count_row}"].border = _border()
    ws[f"D{inv_count_row}"].value = f"=COUNTA(A{inv_data_start}:A{inv_count_row - 1})"
    ws[f"D{inv_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"D{inv_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"D{inv_count_row}"].alignment = _align("center")
    ws[f"D{inv_count_row}"].border = _border()
    ws.row_dimensions[inv_count_row].height = 22

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 4 — מודל עסקי
# ══════════════════════════════════════════════════════════════════════════════
def build_sheet4(wb: Workbook):
    ws = wb.create_sheet("📈 מודל עסקי")
    _set_rtl(ws)

    _set_col_width(ws, {
        "A": 28, "B": 35, "C": 35, "D": 22
    })

    ws.merge_cells("A1:D1")
    _hdr_style(ws["A1"], "מודל עסקי ונוכחות דיגיטלית — Citizen Impact",
               bg=DARK_BLUE, fg=WHITE, size=16)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    _hdr_style(ws["A2"], "B2B SaaS · B2C · פרויקטים ממשלתיים · עוגן",
               bg=MED_BLUE, fg=WHITE, size=11)
    ws.row_dimensions[2].height = 24

    # ── B2B / B2C model ──
    _section_title(ws, 4, 1, 4, "💡  מודל ההכנסות")

    _col_hdr(ws["A5"], "ערוץ")
    _col_hdr(ws["B5"], "לקוח / קהל")
    _col_hdr(ws["C5"], "פתרון / ערך")
    _col_hdr(ws["D5"], "מודל תמחור")
    ws.row_dimensions[5].height = 20

    revenue_data = [
        ("B2B SaaS",                "מוסדות פיננסיים, בנקים",    "הפחתת חוב אבוד, שיפור RAROC",      "מנוי חודשי / שנתי"),
        ("B2B — עיריות",            "רשויות מקומיות, ממשל",      "שירותי חוסן פיננסי לתושבים",       "פרויקט + רישיון"),
        ("B2C ישיר",                "אזרחים פרטיים",              "שיקום פיננסי, ייעוץ AI",            "מנוי משתמש"),
        ("ייעוץ ופרויקטים",         "ארגונים, עמותות",            "פרויקטים ממוקדים (עוגן, מלחמה)",   "Fee per project"),
        ("שותפויות אסטרטגיות",      "Mastercard, בנקים",          "הרחבת כיסוי, גישה ללקוחות",        "Revenue share"),
    ]
    rev_start = 6
    for i, (channel, client, value, pricing) in enumerate(revenue_data):
        row = rev_start + i
        _data_cell(ws[f"A{row}"], channel, i, bold=True)
        _data_cell(ws[f"B{row}"], client, i)
        _data_cell(ws[f"C{row}"], value, i)
        _data_cell(ws[f"D{row}"], pricing, i)
        ws.row_dimensions[row].height = 30

    # Count revenue channels formula
    rev_count_row = rev_start + len(revenue_data)
    ws.merge_cells(f"A{rev_count_row}:C{rev_count_row}")
    ws[f"A{rev_count_row}"].value = "סה\"כ ערוצי הכנסה (נוסחה)"
    ws[f"A{rev_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{rev_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"A{rev_count_row}"].alignment = _align("center")
    ws[f"A{rev_count_row}"].border = _border()
    ws[f"D{rev_count_row}"].value = f"=COUNTA(A{rev_start}:A{rev_count_row - 1})"
    ws[f"D{rev_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"D{rev_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"D{rev_count_row}"].alignment = _align("center")
    ws[f"D{rev_count_row}"].border = _border()
    ws.row_dimensions[rev_count_row].height = 22

    # ── Active Projects ──
    proj_title_row = rev_count_row + 2
    _section_title(ws, proj_title_row, 1, 4, "🚀  פרויקטים פעילים")

    proj_hdr_row = proj_title_row + 1
    _col_hdr(ws[f"A{proj_hdr_row}"], "פרויקט")
    _col_hdr(ws[f"B{proj_hdr_row}"], "קהל יעד")
    _col_hdr(ws[f"C{proj_hdr_row}"], "URL")
    _col_hdr(ws[f"D{proj_hdr_row}"], "שותף")
    ws.row_dimensions[proj_hdr_row].height = 20

    projects = [
        ("לקוחות פרטיים",      "אזרחים פרטיים",       "client.is.citizen-ai.org/#/?customer=35",  "—"),
        ("ייעוץ משכנתאות",     "רוכשי דירות",         "client.is.citizen-ai.org/#/?customer=42",  "—"),
        ("נפגעי מלחמה",        "עוטף עזה + צפון",     "ogen.org",                                  "עוגן"),
        ("CustomerHome 44–48", "סגמנטים שונים",       "landing.is.citizen-ai.org",                "—"),
        ("B2B — בנקים",        "מוסדות פיננסיים",     "www.citizens-ai.com",                       "Mastercard"),
    ]
    proj_data_start = proj_hdr_row + 1
    for i, (proj, audience, url, partner) in enumerate(projects):
        row = proj_data_start + i
        _data_cell(ws[f"A{row}"], proj, i, bold=True)
        _data_cell(ws[f"B{row}"], audience, i)
        _data_cell(ws[f"C{row}"], url, i)
        _data_cell(ws[f"D{row}"], partner, i)
        ws.row_dimensions[row].height = 22

    # Count projects formula
    proj_count_row = proj_data_start + len(projects)
    ws.merge_cells(f"A{proj_count_row}:C{proj_count_row}")
    ws[f"A{proj_count_row}"].value = "סה\"כ פרויקטים פעילים (נוסחה)"
    ws[f"A{proj_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{proj_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"A{proj_count_row}"].alignment = _align("center")
    ws[f"A{proj_count_row}"].border = _border()
    ws[f"D{proj_count_row}"].value = f"=COUNTA(A{proj_data_start}:A{proj_count_row - 1})"
    ws[f"D{proj_count_row}"].font = _font(bold=True, color=WHITE)
    ws[f"D{proj_count_row}"].fill = _fill(DARK_BLUE)
    ws[f"D{proj_count_row}"].alignment = _align("center")
    ws[f"D{proj_count_row}"].border = _border()
    ws.row_dimensions[proj_count_row].height = 22

    # ── Challenges solved ──
    chal_title_row = proj_count_row + 2
    _section_title(ws, chal_title_row, 1, 4, "🎯  אתגרים שנפתרים ללקוחות B2B")

    chal_hdr_row = chal_title_row + 1
    _col_hdr(ws[f"A{chal_hdr_row}"], "אתגר")
    _col_hdr(ws[f"B{chal_hdr_row}"], "פתרון Citizen Impact")
    ws.merge_cells(f"B{chal_hdr_row}:D{chal_hdr_row}")
    ws.row_dimensions[chal_hdr_row].height = 20

    challenges = [
        ("הפחתת חוב אבוד (Bad Debt)", "זיהוי לקוחות בסיכון בזמן אמת + תוכנית שיקום מותאמת"),
        ("שיפור RAROC",               "הגדלת תשואה מסוכנת — לקוחות משקמים הלוואות"),
        ("נאמנות לקוחות",             "ליווי מתמשך שמגדיל Lifetime Value ומוריד Churn"),
        ("עלות ייעוץ גבוהה",          "AI מחליף חלק מהייעוץ האנושי — חיסכון תפעולי"),
    ]
    chal_data_start = chal_hdr_row + 1
    for i, (chal, sol) in enumerate(challenges):
        row = chal_data_start + i
        _data_cell(ws[f"A{row}"], chal, i, bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        _data_cell(ws[f"B{row}"], sol, i)
        ws.row_dimensions[row].height = 26

    # ── Digital presence summary ──
    digital_title_row = chal_data_start + len(challenges) + 1
    _section_title(ws, digital_title_row, 1, 4, "🌐  נוכחות דיגיטלית")

    dig_hdr_row = digital_title_row + 1
    _col_hdr(ws[f"A{dig_hdr_row}"], "פלטפורמה")
    _col_hdr(ws[f"B{dig_hdr_row}"], "כתובת")
    _col_hdr(ws[f"C{dig_hdr_row}"], "קהל")
    _col_hdr(ws[f"D{dig_hdr_row}"], "מטרה")
    ws.row_dimensions[dig_hdr_row].height = 20

    digital = [
        ("אתר B2B",             "www.citizens-ai.com",          "מוסדות פיננסיים", "הצגת הפתרון, מכירה"),
        ("פלטפורמת לקוחות",    "landing.is.citizen-ai.org",    "אזרחים פרטיים",   "גיוס ושימור לקוחות"),
        ("אימייל",              "info@citizens-ai.com",         "כל קהל",          "יצירת קשר"),
        ("טלפון",               "074-7281259",                  "כל קהל",          "תמיכה"),
    ]
    dig_data_start = dig_hdr_row + 1
    for i, (platform, url, audience, purpose) in enumerate(digital):
        row = dig_data_start + i
        _data_cell(ws[f"A{row}"], platform, i, bold=True)
        _data_cell(ws[f"B{row}"], url, i)
        _data_cell(ws[f"C{row}"], audience, i)
        _data_cell(ws[f"D{row}"], purpose, i)
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════
def main():
    output_path = "/home/user/whatsapp-scheduler-v2/citizen_impact_full.xlsx"
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    print("Building Sheet 1: סיכום כללי...")
    build_sheet1(wb)
    print("Building Sheet 2: מוצר וטכנולוגיה...")
    build_sheet2(wb)
    print("Building Sheet 3: צוות ומשקיעים...")
    build_sheet3(wb)
    print("Building Sheet 4: מודל עסקי...")
    build_sheet4(wb)

    print(f"Saving to {output_path}...")
    wb.save(output_path)
    print("Done.")

if __name__ == "__main__":
    main()
