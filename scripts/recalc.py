"""
recalc.py — validates an openpyxl workbook and reports any issues.
Usage: python scripts/recalc.py <path_to_xlsx>
Output: JSON with "status" and "total_errors".
"""
import sys
import json
from openpyxl import load_workbook


def recalc(path: str) -> None:
    sheet_info = []
    errors = []

    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        result = {
            "status": "error",
            "total_errors": 1,
            "errors": [f"Cannot open {path}: {e}"],
            "sheets": [],
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(1)

    for name in wb.sheetnames:
        ws = wb[name]
        sheet_info.append({
            "name": name,
            "dims": ws.dimensions,
            "rtl": ws.sheet_view.rightToLeft,
        })
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    if formula.count("(") != formula.count(")"):
                        errors.append(
                            f"{name}!{cell.coordinate}: unbalanced parens in {formula!r}"
                        )

    result = {
        "status": "success" if not errors else "error",
        "total_errors": len(errors),
        "sheets": sheet_info,
    }
    if errors:
        result["errors"] = errors

    print(json.dumps(result, ensure_ascii=False, indent=2))
    if errors:
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "total_errors": 1,
                          "errors": ["Usage: python scripts/recalc.py <path_to_xlsx>"]},
                         ensure_ascii=False, indent=2))
        sys.exit(1)
    recalc(sys.argv[1])
