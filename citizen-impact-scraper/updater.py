"""
updater.py — opens citizen_impact_full.xlsx and updates cells with
live data from ScraperResult. Only touches value cells; never overwrites
Excel formulas.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scraper import ScraperResult, PageInfo

# Cell map: sheet name → { field: (row, col) }
# Matches the layout built by scripts/build_citizen_impact.py

SHEET_SUMMARY  = "📋 סיכום כללי"
SHEET_PRODUCT  = "🛠️ מוצר וטכנולוגיה"

# Contact cells in SHEET_SUMMARY  (row, col) — 1-indexed
CONTACT_EMAIL_CELL = (13, 2)  # B13
CONTACT_PHONE_CELL = (14, 2)  # B14

# KPI base-value cells in SHEET_SUMMARY  (col B, rows 23–26)
KPI_CELLS = [
    (23, 2),  # B23 — גידול הכנסות (משתמשים %)
    (24, 2),  # B24 — גידול הכנסות (שיפור %)
    (25, 2),  # B25 — מגירעון ליתרת זכות
    (26, 2),  # B26 — כשירות אשראי
]
KPI_LABELS = [
    "גידול בהכנסות",
    "גידול בהכנסות",
    "מגירעון ליתרת זכות",
    "כשירות אשראי",
]

# Pages table in SHEET_PRODUCT starts at row 22, cols B–D
PAGES_START_ROW = 22
PAGES_COL_NAME  = 2  # B
PAGES_COL_URL   = 3  # C
PAGES_COL_STATUS = 4  # D


def _thin_border() -> Border:
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _row_fill(i: int) -> PatternFill:
    return _fill("DBEAFE" if i % 2 == 0 else "FFFFFF")


def _data_style(cell, i: int, bold: bool = False) -> None:
    cell.fill = _row_fill(i)
    cell.font = Font(name="Arial", size=10, bold=bold, color="1E293B")
    cell.alignment = Alignment(horizontal="right", vertical="center",
                               wrap_text=True, readingOrder=2)
    cell.border = _thin_border()


def update_contact(ws, result: ScraperResult) -> int:
    """Updates email and phone cells. Returns number of cells changed."""
    changed = 0
    if result.contact.email:
        r, c = CONTACT_EMAIL_CELL
        if ws.cell(r, c).value != result.contact.email:
            ws.cell(r, c).value = result.contact.email
            changed += 1
    if result.contact.phone:
        r, c = CONTACT_PHONE_CELL
        if ws.cell(r, c).value != result.contact.phone:
            ws.cell(r, c).value = result.contact.phone
            changed += 1
    return changed


def update_kpis(ws, result: ScraperResult) -> int:
    """Updates KPI base values. Formulas in col C/D recalculate automatically."""
    if not result.kpis:
        return 0

    kpi_map: dict[str, int] = {k.label: k.value for k in result.kpis}
    changed = 0
    for (row, col), label in zip(KPI_CELLS, KPI_LABELS):
        if label in kpi_map:
            new_val = kpi_map[label]
            cell = ws.cell(row, col)
            if cell.value != new_val:
                cell.value = new_val
                changed += 1
    return changed


def update_pages(ws, pages: list[PageInfo]) -> int:
    """
    Replaces the pages data block in SHEET_PRODUCT.
    Only touches rows PAGES_START_ROW … PAGES_START_ROW+len(pages)-1.
    Does NOT touch the COUNTA formula row below.
    """
    if not pages:
        return 0

    changed = 0
    for i, page in enumerate(pages):
        row = PAGES_START_ROW + i
        name_cell   = ws.cell(row, PAGES_COL_NAME)
        url_cell    = ws.cell(row, PAGES_COL_URL)
        status_cell = ws.cell(row, PAGES_COL_STATUS)

        if name_cell.value != page.name:
            name_cell.value = page.name
            _data_style(name_cell, i, bold=True)
            changed += 1
        if url_cell.value != page.url:
            url_cell.value = page.url
            _data_style(url_cell, i)
            changed += 1
        if status_cell.value != page.status:
            status_cell.value = page.status
            _data_style(status_cell, i, bold=True)
            status_cell.font = Font(name="Arial", size=10, bold=True, color="10B981")
            changed += 1

    return changed


def run_update(excel_path: str, result: ScraperResult) -> dict:
    path = Path(excel_path)
    if not path.exists():
        return {"ok": False, "error": f"File not found: {excel_path}"}

    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}

    summary = {}

    if SHEET_SUMMARY in wb.sheetnames:
        ws1 = wb[SHEET_SUMMARY]
        summary["contact_cells_updated"] = update_contact(ws1, result)
        summary["kpi_cells_updated"]     = update_kpis(ws1, result)
    else:
        summary["contact_cells_updated"] = 0
        summary["kpi_cells_updated"]     = 0

    if SHEET_PRODUCT in wb.sheetnames:
        ws2 = wb[SHEET_PRODUCT]
        summary["page_cells_updated"] = update_pages(ws2, result.pages)
    else:
        summary["page_cells_updated"] = 0

    try:
        wb.save(path)
        summary["ok"] = True
        summary["saved_to"] = str(path.resolve())
    except Exception as exc:
        summary["ok"] = False
        summary["error"] = str(exc)

    return summary


if __name__ == "__main__":
    import json
    from scraper import scrape_all
    import config

    print("Running scraper...")
    result = scrape_all()

    print("Running updater...")
    summary = run_update(config.EXCEL_PATH, result)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    sys.exit(0 if summary.get("ok") else 1)
